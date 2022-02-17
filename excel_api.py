import os
import openpyxl as pyxl
import shutil
import json


class ExcelAPI:
    def __init__(self):
        """
        The __init__ function is the constructor for a class. It initializes the attributes of an object.
        :param self: Used to refer to the object itself.
        :return: None.
        """
        self.workbook = None
        self.point_dicts = {}
        self.path_to_file = None
        self.point_sheet_types = [""]
        self.name_list = []
    
    def open_existing_file(self, path_to_file):
        """
        The open_existing_file function opens an existing file and creates a dictionary of the points in that file.
        It also sets the path_to_file variable to be equal to the path passed into it.
        
        :param self: Used to reference the object itself.
        :param path_to_file: Used to specify the path to the file that is being opened.
        :return: True if operation was successfull
        """
        if path_to_file == "":
            return False

        self.path_to_file = path_to_file
        self.workbook = pyxl.load_workbook(filename=self.path_to_file)
        self.create_name_list()
        self.create_point_dicts()
        return True
    
    def create_point_dicts(self):
        """
        The create_point_dicts function creates a dictionary for each point sheet type in the workbook. 
        
        :param self: Used to access the instance variables of the class.
        """
        for point_type in list(set(self.point_sheet_types)):
            point_type_name = "".join(["Punkte", point_type])
            point_sheet = self.workbook[str(point_type_name)]
            self.point_dicts[point_type_name] = self.point_map_dict(point_sheet)
    
    def create_name_list(self):
        """
        The create_name_list function creates a list of all the names of the attending students in the excel file. 
        
        :param self: Used to access the class attributes.
        """
        name_sheet = self.workbook["Namensliste"]
        for row in name_sheet.iter_rows(min_row=2,
                                max_row=500,
                                min_col=2,
                                max_col=6,
                                values_only=True):
            if row[0] != None or row[2] != None:
                if row[4] != None:
                    self.point_sheet_types.append(row[4].upper())
                turn_up = str(row[3]).replace(".", "").replace(",","")
                turn_up = turn_up.lower()
                if turn_up == "ne":
                    pass
                else:
                    if row[4] == None:
                        examtype = ""
                    else:
                        examtype = row[4]
                    self.name_list.append([str(row[0]), str(row[2]), str(examtype)])
    
    @staticmethod
    def point_map_dict(point_sheet):
        point_dict = {}
        character_arr = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
        task_list = ["Task 1", "Task 2", "Task 3", "Task 4", "Task 5", "Task 6", "Task 7", "Task 8", "Task 9", "Task 10"]
        for task_ind, task_name in enumerate(task_list):
            comb_columns = []
            point_dict[task_name] = {
                "A": {},
                "B": {},
                "C": {},
                "D": {},
                "E": {},
                "F": {},
                "G": {},
                "H": {},
                "I": {},
                "J": {},
                "K": {},
                "L": {},
                "M": {},
                "N": {},
                "O": {},
                "P": {},
                "Q": {},
                "R": {},
                "S": {},
                "T": {},
                "U": {},
                "V": {},
                "W": {},
                "X": {},
                "Y": {},
                "Z": {},
            }
            for column in point_sheet.iter_cols(min_row=task_ind*9+3,
                                                max_row=task_ind*9+5,
                                                min_col=1,
                                                max_col=13,
                                                values_only=True):
                comb_columns.append(column)
            for column in point_sheet.iter_cols(min_row=task_ind*9+7,
                                                max_row=task_ind*9+9,
                                                min_col=1,
                                                max_col=13,
                                                values_only=True):
                comb_columns.append(column)
            for arr_ind, value in enumerate(comb_columns):
                point_dict[task_name][character_arr[arr_ind]] = {"P": value[0], "V":value[1], "U":value[2]}
        return point_dict
    
    @staticmethod
    def combine_marked_point_dicts(point_dict, filled_dict):
        """
        The combine_marked_point_dicts function combines the point_dict and filled_dict into one dictionary, 
        which is then used to calculate the total points for each task.
        
        :param point_dict: Used to store the valid answers of the exam.
        :param filled_dict: Used to store the marked answers by the student.
        :return: A combined dict of the input dictionaries.
        """
        for task in point_dict:
            for char in point_dict[task]:
                if char in filled_dict[task]:
                    point_dict[task][char]["M"] = 1
                else:
                    point_dict[task][char]["M"] = 0
        return point_dict

    @staticmethod
    def point_dict_to_subtasks(point_dict):
        """
        The point_dict_to_subtasks function splits the point_dict into it's subtasks.
        
        :param point_dict: Combined point_dict from the combine_marked_point_dicts.
        :return: The point_dict split into its subtasks.
        """
        complete_sub_task_dict = {}
        for task_ind, task in enumerate(point_dict):
            sub_task_dict = {}
            last_sub_number = 1
            for char in point_dict[task]:
                if point_dict[task][char]["U"] > last_sub_number:
                    complete_sub_task_dict["Aufgabe "+str(task_ind+1)+"."+str(point_dict[task][char]["U"]-1)] = sub_task_dict
                    sub_task_dict = {}
                    last_sub_number = point_dict[task][char]["U"]
                if point_dict[task][char]["U"] != 0:
                    sub_task_dict[char] =  point_dict[task][char]
            complete_sub_task_dict["Aufgabe "+str(task_ind+1)+"."+str(last_sub_number)] = sub_task_dict
        return complete_sub_task_dict
    
    def copy_template(self, mat_number):
        try:
            self.workbook[mat_number]
        except KeyError:
            template_sheet = self.workbook["Vorlage"]
            self.workbook.copy_worksheet(template_sheet)
            new_sheet = self.workbook[self.workbook.sheetnames[-1]]
            new_sheet.title = mat_number

    def fill_sheet(self, mat_number, subtask_dict):
        character_arr = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
        sheet = self.workbook[mat_number]
        subtask_row = 14
        total_possible_points = 0
        total_achieved_points = 0
        for sub_task in subtask_dict:
            sub_task_result = 0
            char_start = 4
            sheet["A"+str(subtask_row)] = sub_task
            sheet["A"+str(subtask_row)].font = pyxl.styles.Font(bold=True)

            for char in subtask_dict[sub_task]:
                sheet[character_arr[char_start]+str(subtask_row)] = char
                sheet[character_arr[char_start]+str(subtask_row)].font = pyxl.styles.Font(bold=True)
                sheet[character_arr[char_start]+str(subtask_row)].alignment =  pyxl.styles.Alignment(horizontal="right")
                
                sheet[character_arr[char_start]+str(subtask_row+1)] = subtask_dict[sub_task][char]["M"]
                sub_task_result += subtask_dict[sub_task][char]["M"] * subtask_dict[sub_task][char]["P"]
                char_start += 1

            sheet["B"+str(subtask_row+1)] = sub_task_result
            sheet["C"+str(subtask_row+1)] = max(sub_task_result, 0)
            sheet["C"+str(subtask_row+1)].font = pyxl.styles.Font(bold=True)
            sheet["D"+str(subtask_row+1)] = "von " + str(abs(subtask_dict[sub_task][char]["P"]))

            total_possible_points += abs(subtask_dict[sub_task][char]["P"])
            total_achieved_points += max(sub_task_result, 0)


            subtask_row += 3
            sheet["D"+str(9)] = total_achieved_points
            sheet["D"+str(10)] = total_possible_points
            sheet["D"+str(11)] = total_achieved_points/total_possible_points
    
    def add_personal_data(self, name, filename, mat_number, examtype):
        sheet = self.workbook[str(mat_number)]
        sheet["B"+str(6)] = name
        sheet["I"+str(6)] = filename
        sheet["L"+str(6)] = mat_number
        sheet["R"+str(6)] = examtype
    
    def save_in_sheet(self, mat_number, name, marked_dict, filename, examtype):
        self.copy_template(str(mat_number))
        point_type_name = "".join(["Punkte", examtype])
        combinded_dict = self.combine_marked_point_dicts(self.point_dicts[point_type_name], marked_dict)
        subtask_dict = self.point_dict_to_subtasks(combinded_dict)
        self.fill_sheet(str(mat_number), subtask_dict)
        self.add_personal_data(name, filename, str(mat_number), examtype)
        self.workbook.save(filename=self.path_to_file)

if __name__ == "__main__":
    excel_api = ExcelAPI()
    excel_api.open_existing_file("excel_shit/NachAuswertung.xlsx")
